import io
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, Series
import pandas as pd

DIVISIONS = [
    "Denver", "Haggen", "Intermountain", "Jewel-Osco", "MidAtlantic",
    "NorCal", "Portland", "Seattle", "Shaws", "SoCal", "Southern",
    "Southwest", "United"
]
TYPES = ["SYSTEM", "OPS", "PE"]

# ── Styles ─────────────────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="4F46E5")   # indigo
HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
SUB_FILL   = PatternFill("solid", fgColor="E0E7FF")   # light indigo
SUB_FONT   = Font(bold=True, color="1E1B4B", size=10)
BOLD_FONT  = Font(bold=True, size=10)
NORM_FONT  = Font(size=10)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CTR_ALIGN  = Alignment(horizontal="center", vertical="center")
TOP_LEFT   = Alignment(vertical="top")

def _thin_border():
    s = Side(border_style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)

def _apply_header(ws, row_num, values, fill, font):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=col, value=val)
        c.fill   = fill
        c.font   = font
        c.border = _thin_border()
        c.alignment = CTR_ALIGN

def _auto_width(ws, min_w=8, max_w=40):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        best = min_w
        for cell in col:
            if cell.value:
                lines = str(cell.value).split("\n")
                best = max(best, max(len(l) for l in lines))
        ws.column_dimensions[col_letter].width = min(best + 2, max_w)

def _freeze(ws, cell="B3"):
    ws.freeze_panes = cell


# ── Tab builders ────────────────────────────────────────────────────────────

def write_inc(ws_or_wb, inc: pd.DataFrame, sheet_name: str = "INC"):
    if isinstance(ws_or_wb, Workbook):
        ws = ws_or_wb.create_sheet(sheet_name)
    else:
        ws = ws_or_wb
    cols = list(inc.columns)
    _apply_header(ws, 1, cols, HDR_FILL, HDR_FONT)
    for r_idx, row in enumerate(inc.itertuples(index=False), 2):
        for c_idx, val in enumerate(row, 1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.font      = NORM_FONT
            c.border    = _thin_border()
            c.alignment = WRAP_ALIGN
    _auto_width(ws)
    _freeze(ws, "A2")
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"


def _write_category_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame, value_col: str = "count"):
    """
    Shared layout for Acupick, Acupick Incident Categories, and Ops Issues.
    value_col: 'count' writes numbers; 'caller' writes caller strings.
    """
    ws = wb.create_sheet(sheet_name)

    # Row 1: section label
    ws.cell(row=1, column=1, value="")
    ws.cell(row=1, column=2, value="Operations Issue").font = BOLD_FONT
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=2 + len(DIVISIONS))

    # Row 2: column headers
    headers = ["", "Category"] + DIVISIONS + ["Total"]
    _apply_header(ws, 2, headers, SUB_FILL, SUB_FONT)

    # Data rows
    for r_idx, row in enumerate(df.itertuples(index=False), 3):
        # Code cell
        code_cell = ws.cell(row=r_idx, column=1, value=row.Code)
        code_cell.font      = BOLD_FONT
        code_cell.border    = _thin_border()
        code_cell.alignment = CTR_ALIGN

        # Category label
        lbl = ws.cell(row=r_idx, column=2, value=row.Category)
        lbl.font      = NORM_FONT
        lbl.border    = _thin_border()
        lbl.alignment = WRAP_ALIGN

        # Division values
        for c_idx, div in enumerate(DIVISIONS, 3):
            val = getattr(row, div, 0)
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font      = NORM_FONT
            cell.border    = _thin_border()
            cell.alignment = WRAP_ALIGN if value_col == "caller" else CTR_ALIGN

        # Total
        tot = ws.cell(row=r_idx, column=3 + len(DIVISIONS), value=row.Total)
        tot.font      = BOLD_FONT
        tot.border    = _thin_border()
        tot.alignment = CTR_ALIGN

    _auto_width(ws)
    _freeze(ws, "C3")


def write_acupick_categories(wb: Workbook, categories: dict):
    ws = wb.create_sheet("Acupick Incident Categories")
    TYPES_LABELS = {"OPS": "Operations Issue", "SYSTEM": "System Issue", "PE": "Product Enhancement"}
    row_num = 1

    for type_key, type_label in TYPES_LABELS.items():
        df = categories.get(type_key, pd.DataFrame())
        if df is None or len(df) == 0:
            continue

        # Section header row
        ws.cell(row=row_num, column=1, value="")
        label_cell = ws.cell(row=row_num, column=2, value=type_label)
        label_cell.font = BOLD_FONT
        label_cell.fill = HDR_FILL
        label_cell.font = HDR_FONT
        ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=2 + len(DIVISIONS))
        row_num += 1

        # Column headers
        headers = ["", "Category"] + DIVISIONS + ["Total"]
        for c_idx, val in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=c_idx, value=val)
            cell.fill = SUB_FILL; cell.font = SUB_FONT
            cell.border = _thin_border(); cell.alignment = CTR_ALIGN
        row_num += 1

        # Data rows
        summary_row = {d: 0 for d in DIVISIONS}
        summary_row["Total"] = 0
        for _, r in df.iterrows():
            ws.cell(row=row_num, column=1, value=r["Code"]).font = BOLD_FONT
            ws.cell(row=row_num, column=1).border = _thin_border()
            ws.cell(row=row_num, column=1).alignment = CTR_ALIGN
            ws.cell(row=row_num, column=2, value=r["Category"]).font = NORM_FONT
            ws.cell(row=row_num, column=2).border = _thin_border()
            for c_idx, div in enumerate(DIVISIONS, 3):
                val = int(r.get(div, 0))
                cell = ws.cell(row=row_num, column=c_idx, value=val)
                cell.font = NORM_FONT; cell.border = _thin_border(); cell.alignment = CTR_ALIGN
                summary_row[div] += val
            tot = int(r.get("Total", 0))
            ws.cell(row=row_num, column=3+len(DIVISIONS), value=tot).font = BOLD_FONT
            ws.cell(row=row_num, column=3+len(DIVISIONS)).border = _thin_border()
            ws.cell(row=row_num, column=3+len(DIVISIONS)).alignment = CTR_ALIGN
            summary_row["Total"] += tot
            row_num += 1

        # Summary row
        ws.cell(row=row_num, column=1, value="")
        ws.cell(row=row_num, column=2, value="Summary").font = BOLD_FONT
        ws.cell(row=row_num, column=2).fill = SUB_FILL
        ws.cell(row=row_num, column=2).border = _thin_border()
        for c_idx, div in enumerate(DIVISIONS, 3):
            cell = ws.cell(row=row_num, column=c_idx, value=summary_row[div])
            cell.font = BOLD_FONT; cell.fill = SUB_FILL
            cell.border = _thin_border(); cell.alignment = CTR_ALIGN
        tot_cell = ws.cell(row=row_num, column=3+len(DIVISIONS), value=summary_row["Total"])
        tot_cell.font = BOLD_FONT; tot_cell.fill = SUB_FILL
        tot_cell.border = _thin_border(); tot_cell.alignment = CTR_ALIGN
        row_num += 2  # blank row between sections

    _auto_width(ws)
    _freeze(ws, "C3")


def write_acupick(wb: Workbook, df: pd.DataFrame):
    _write_category_sheet(wb, "Acupick", df, value_col="count")


def write_ops_issues(wb: Workbook, df: pd.DataFrame):
    _write_category_sheet(wb, "Ops Issues", df, value_col="caller")


def write_summary(wb: Workbook, summary: dict):
    ws        = wb.create_sheet("Summary")
    date_from = summary["date_from"]
    date_to   = summary["date_to"]
    daily     = summary["daily"]
    totals    = summary["totals"]
    TYPES     = ["SYSTEM", "OPS", "PE"]

    row_num = 1

    # ── Totals section ───────────────────────────────────────────────
    ws.cell(row=row_num, column=1, value=f"{date_from} to {date_to}").font = BOLD_FONT
    for c_idx, hdr in enumerate(TYPES + ["Total"], 2):
        cell = ws.cell(row=row_num, column=c_idx, value=hdr)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.border = _thin_border(); cell.alignment = CTR_ALIGN
    row_num += 1

    for r in totals:
        ws.cell(row=row_num, column=1, value=r["Configuration Item"]).font = NORM_FONT
        ws.cell(row=row_num, column=1).border = _thin_border()
        for c_idx, t in enumerate(TYPES + ["Total"], 2):
            cell = ws.cell(row=row_num, column=c_idx, value=r.get(t, 0))
            cell.font = BOLD_FONT if t == "Total" else NORM_FONT
            cell.border = _thin_border(); cell.alignment = CTR_ALIGN
        row_num += 1

    row_num += 1  # blank separator

    # ── Per-day sections ─────────────────────────────────────────────
    for day_block in daily:
        # Day header row
        ws.cell(row=row_num, column=1, value=day_block["date"]).font = BOLD_FONT
        ws.cell(row=row_num, column=1).fill = SUB_FILL
        for c_idx, hdr in enumerate(TYPES, 2):
            cell = ws.cell(row=row_num, column=c_idx, value=hdr)
            cell.fill = SUB_FILL; cell.font = SUB_FONT
            cell.border = _thin_border(); cell.alignment = CTR_ALIGN
        row_num += 1

        for r in day_block["data"]:
            ws.cell(row=row_num, column=1, value=r["Configuration Item"]).font = NORM_FONT
            ws.cell(row=row_num, column=1).border = _thin_border()
            for c_idx, t in enumerate(TYPES, 2):
                cell = ws.cell(row=row_num, column=c_idx, value=r.get(t, 0))
                cell.font = NORM_FONT; cell.border = _thin_border(); cell.alignment = CTR_ALIGN
            row_num += 1

        row_num += 1  # blank separator between days

    _auto_width(ws)
    _freeze(ws, "B2")


# ── Main entry point ─────────────────────────────────────────────────────────

def write_vpos_categories(wb: Workbook, categories: dict):
    ws = wb.create_sheet("VPOS Incident Categories")
    TYPES_LABELS = {"OPS": "Operations Issue", "SYSTEM": "System Issue", "PE": "Product Enhancement"}
    row_num = 1

    for type_key, type_label in TYPES_LABELS.items():
        df = categories.get(type_key, pd.DataFrame())
        if df is None or len(df) == 0:
            continue

        ws.cell(row=row_num, column=1, value="")
        label_cell = ws.cell(row=row_num, column=2, value=type_label)
        label_cell.fill = HDR_FILL
        label_cell.font = HDR_FONT
        ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=2+len(DIVISIONS))
        row_num += 1

        headers = ["", "Category"] + DIVISIONS + ["Total"]
        for c_idx, val in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=c_idx, value=val)
            cell.fill = SUB_FILL; cell.font = SUB_FONT
            cell.border = _thin_border(); cell.alignment = CTR_ALIGN
        row_num += 1

        summary_row = {d: 0 for d in DIVISIONS}
        summary_row["Total"] = 0
        for _, r in df.iterrows():
            ws.cell(row=row_num, column=1, value=r["Code"]).font = BOLD_FONT
            ws.cell(row=row_num, column=1).border = _thin_border()
            ws.cell(row=row_num, column=1).alignment = CTR_ALIGN
            ws.cell(row=row_num, column=2, value=r["Category"]).font = NORM_FONT
            ws.cell(row=row_num, column=2).border = _thin_border()
            for c_idx, div in enumerate(DIVISIONS, 3):
                val = int(r.get(div, 0))
                cell = ws.cell(row=row_num, column=c_idx, value=val)
                cell.font = NORM_FONT; cell.border = _thin_border(); cell.alignment = CTR_ALIGN
                summary_row[div] += val
            tot = int(r.get("Total", 0))
            ws.cell(row=row_num, column=3+len(DIVISIONS), value=tot).font = BOLD_FONT
            ws.cell(row=row_num, column=3+len(DIVISIONS)).border = _thin_border()
            ws.cell(row=row_num, column=3+len(DIVISIONS)).alignment = CTR_ALIGN
            summary_row["Total"] += tot
            row_num += 1

        ws.cell(row=row_num, column=1, value="")
        ws.cell(row=row_num, column=2, value="Summary").font = BOLD_FONT
        ws.cell(row=row_num, column=2).fill = SUB_FILL
        ws.cell(row=row_num, column=2).border = _thin_border()
        for c_idx, div in enumerate(DIVISIONS, 3):
            cell = ws.cell(row=row_num, column=c_idx, value=summary_row[div])
            cell.font = BOLD_FONT; cell.fill = SUB_FILL
            cell.border = _thin_border(); cell.alignment = CTR_ALIGN
        tot_cell = ws.cell(row=row_num, column=3+len(DIVISIONS), value=summary_row["Total"])
        tot_cell.font = BOLD_FONT; tot_cell.fill = SUB_FILL
        tot_cell.border = _thin_border(); tot_cell.alignment = CTR_ALIGN
        row_num += 2

    _auto_width(ws)
    _freeze(ws, "C3")


def build_workbook(processed: dict, report_type: str = "acupick") -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    if report_type == "vpos":
        vpos_inc = processed["vpos_inc"] if isinstance(processed.get("vpos_inc"), pd.DataFrame) and len(processed["vpos_inc"]) else processed["inc"]
        write_vpos_categories(wb, processed["vpos_categories"])
        write_ops_issues(wb, processed["vpos_ops"])
        write_inc(wb, vpos_inc, "INC")
    else:
        acu_inc = processed["acupick_inc"] if isinstance(processed.get("acupick_inc"), pd.DataFrame) and len(processed["acupick_inc"]) else processed["inc"]
        write_ops_issues(wb, processed["acupick_ops"])
        write_summary(wb, processed["summary"])
        write_acupick_categories(wb, processed["acupick_categories"])
        write_inc(wb, acu_inc, "INC")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()